import xml.etree.ElementTree as ET
import os
import pyodbc
import time
import re
import logging
import argparse
import configparser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from tkinter import filedialog
from tkinter import Tk

IGNORED_VALUES = ['ACTIVE','REVIEW', 'ENDED', 'N/A', '385432009','C','U','R','RD','999011011000230107','12464001000001103']
NAMESPACE = {'ns': 'http://www.e-mis.com/emisopen'}
# Set up argument parsing
parser = argparse.ArgumentParser(description='Process configuration paths.')
parser.add_argument('--xml_directory', type=str)
parser.add_argument('--database_path', type=str)
parser.add_argument('--transitive_closure_db_path', type=str)
parser.add_argument('--history_db_path', type=str)
parser.add_argument('--output_dir', type=str)

args = parser.parse_args()

# Load config values from config.ini if arguments are not provided
config = configparser.ConfigParser()
config.read('config.ini')

xml_directory = args.xml_directory if args.xml_directory else config['DEFAULT']['xml_directory']
database_path = args.database_path if args.database_path else config['DEFAULT']['database_path']
transitive_closure_db_path = args.transitive_closure_db_path if args.transitive_closure_db_path else config['DEFAULT']['transitive_closure_db_path']
history_db_path = args.history_db_path if args.history_db_path else config['DEFAULT']['history_db_path']
output_dir = args.output_dir if args.output_dir else config['DEFAULT']['output_dir']

start_time = time.time() #start clock

# Set up logging
log_filename = os.path.join(output_dir, "log.txt")
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

# Clear the log file
with open(log_filename, 'w'):
    pass

# After setting up basicConfig
console = logging.StreamHandler()
console.setLevel(logging.INFO)
logging.getLogger('').addHandler(console)

# List and log XML files
xml_files = [f for f in os.listdir(xml_directory) if f.endswith('.xml')]
logging.info(f"Found {len(xml_files)} XML files in the directory: {xml_directory}")
for file in xml_files:
    logging.info(f"{file}")
logging.info("-" * 40)

def sanitize_filename(filename):
    """Remove characters that are illegal in filenames on Windows."""
    illegal_chars = r'<>:"/\|?*'
    for char in illegal_chars:
        filename = filename.replace(char, '_')
    return filename

def process_single_report(data, report_name, database_path, transitive_closure_db_path, output_dir):
    # Generate the output file path based on the report name
    output_file = os.path.join(output_dir, f"snomed_codes_{report_name}.xlsx")
    
    # Create output folder if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Save the report data to the Excel file
    processed_value_sets, total_value_sets = save_to_xlsx(data, output_file, connection_main, connection_tc, connection_history)
    
    if processed_value_sets > 0:
        logging.info(f"For report '{report_name}', successfully processed {processed_value_sets}/{total_value_sets} value sets.")
        logging.info(f"Excel workbook saved to {output_file}\n")
    else:
        logging.info(f"For report '{report_name}', no value sets were processed as they didn't contain any SNOMED-CT Concepts. No workbook saved.\n")

def extract_and_process_reports_from_xml(xml_path, database_path, transitive_closure_db_path, output_dir):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    reports = root.findall(".//ns:report", NAMESPACE)
    logging.info(f"Found {len(reports)} reports:")

    # Print names of all reports
    report_names = [report.find(".//ns:name", NAMESPACE).text for report in reports]
    for name in report_names:
        logging.info(name)
    logging.info("-" * 40)  # Print separator for clarity

    for report in reports:
        report_name = report.find(".//ns:name", NAMESPACE).text 
        logging.info(f"Processing report: {report_name}")
        name_content = report.find("ns:name", NAMESPACE).text
        match = re.search(r'\[(.*?)\]', name_content)
        report_name_sanitized = sanitize_filename(report_name)  # Sanitize the report name here
        report_name = match.group(1) if match else report_name_sanitized
        extracted_data = extract_values_from_xml_element(report)
        process_single_report(extracted_data, report_name, database_path, transitive_closure_db_path, output_dir)

def extract_values_from_xml_element(element):
    data_sets = []
    seen_data_sets = set()

    for valueSet in element.findall(".//ns:valueSet", NAMESPACE):  # Change criterion to valueSet
        data = []

        for values in valueSet.findall(".//ns:values", NAMESPACE):
            value_elem = values.find('ns:value', NAMESPACE)
            value = value_elem.text if value_elem is not None else "N/A"

            if value in IGNORED_VALUES:
                continue

            displayName_elem = values.find('ns:displayName', NAMESPACE)
            displayName = displayName_elem.text if displayName_elem is not None else "N/A"

            includeChildren_elem = values.find('ns:includeChildren', NAMESPACE)
            includeChildren = includeChildren_elem.text if includeChildren_elem is not None else "false"

            exception_codes = set()
            for exception in valueSet.findall('.//ns:exception//ns:values/ns:value', NAMESPACE):  # Updated XPath
                exception_codes.add(exception.text) 

            data.append((value, displayName, includeChildren, frozenset(exception_codes)))

        tuple_data = tuple(data)
        if tuple_data not in seen_data_sets and len(data) > 0:
            seen_data_sets.add(tuple_data)
            data_sets.append(data)
    
    logging.info(f"Extracted {len(data_sets)} datasets from the XML.")
    for i, dataset in enumerate(data_sets, 1):
        logging.info(f"Dataset {i}/{len(data_sets)} contains {len(dataset)} value{'s' if len(dataset) != 1 else ''}.")

    return [list(ds) for ds in data_sets]

def get_cui_from_access(tui_list, display_names, connection_main):
    tui_to_cui = {}
    display_name_to_cui = {}
    logging.info(f"Using the established connection to the database")

    # Ensure that the TUI list has distinct values
    distinct_tui_list = list(set(tui_list))

    # Query the database for CUI values based on the provided TUI list
    if distinct_tui_list:
        query_for_tui = "SELECT TUI, CUI FROM SCT WHERE TUI IN ({})".format(','.join(['?'] * len(distinct_tui_list)))
        cursor = connection_main.cursor()
        cursor.execute(query_for_tui, distinct_tui_list)
        tui_to_cui = {row.TUI: row.CUI for row in cursor.fetchall()}
        logging.info(f"Queried {len(tui_to_cui)} Description IDs from XML with SNOMED Description IDs (TUI) from the database and return Concept IDs (CUI).")

    # Query the database for CUI values based on the provided DisplayNames (Terms)
    if display_names: 
        query_for_display_name = "SELECT Term, CUI FROM SCT WHERE Term IN ({})".format(','.join(['?'] * len(display_names)))
        cursor = connection_main.cursor()
        cursor.execute(query_for_display_name, display_names)
        display_name_to_cui = {row.Term: row.CUI for row in cursor.fetchall()}
        logging.info(f"Queried {len(display_name_to_cui)} DisplayNames from XML with SNOMED Description IDs (TUI) from the database and return Concept IDs (CUI).")

    return tui_to_cui, display_name_to_cui

def get_all_children_from_database(code, connection, exceptions=None):
    if exceptions is None:
        exceptions = set()
    
    children = {code}
    newly_added = {code}
    
    loop_count = 0

    while newly_added:
        # Increment the loop count
        loop_count += 1

        # Get direct children of the codes in newly_added
        query = "SELECT SubtypeID FROM SCTTC WHERE SupertypeID IN ({})".format(",".join(['?'] * len(newly_added)))
        cursor = connection.cursor()
        cursor.execute(query, list(newly_added))
        new_children = {row.SubtypeID for row in cursor.fetchall()}
        
        # Remove the ones we already know about to avoid infinite loops
        newly_added = new_children - children

        # Exclude exception codes
        excluded_codes = newly_added.intersection(exceptions)
        if excluded_codes:
            logging.info(f"Excluded child codes for {code}: {', '.join(map(str, excluded_codes))}")
        newly_added -= exceptions

        # Add the newly discovered children to main set
        children.update(newly_added)

    if len(children) == 1:  # Only the code itself, no children
        additional_msg = ". No child codes found."
    elif len(children) > 1:  # Code itself plus additional children
        additional_msg = f". Found {len(children) - 1} child codes."
    else:  # This situation shouldn't occur, but we'll handle it just in case
        additional_msg = ""

    logging.info(f"Fetching children for code {code} completed in {loop_count} iteration{'s' if loop_count > 1 else ''}{additional_msg}")
    return children

def get_new_cui_from_history(old_cui_list, connection_history):
    if not old_cui_list:  # Check if the list is empty
        return {}
    cursor = connection_history.cursor()
    placeholders = ', '.join('?' for _ in old_cui_list)
    query = f"SELECT OLDCUI, NEWCUI FROM SCTHIST WHERE OLDCUI IN ({placeholders})"
    cursor.execute(query, old_cui_list)
    results = {row.OLDCUI: row.NEWCUI for row in cursor.fetchall()}
    
    # Print details about the history lookups
    for old_cui, new_cui in results.items():
        if new_cui:
            logging.info(f"Old Concept ID: {old_cui} has a new Concept ID: {new_cui}")
        else:
            logging.info(f"No new Concept ID found for: {old_cui}")

    # Ensure that all OLDCUI values are in the results dictionary, with a value of None if no match was found
    for old_cui in old_cui_list:
        results.setdefault(old_cui, None)
    
    return results

def save_to_xlsx(data, file_path, connection_main, connection_tc, connection_history):
    total_value_sets = len(data)
    logging.info(f"Starting to process {total_value_sets} data sets, lookup concept IDs and identify any relevant child codes")
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)  # Remove the default sheet created

    # Initialize a set to keep track of all Concept IDs that have been checked
    checked_cuis = set()

    for idx, value_set_data in enumerate(data, 1):
        logging.info(f"Processing value set {idx}/{total_value_sets}:")
        tui_values = [entry[0] for entry in value_set_data]
        display_names = [entry[1] for entry in value_set_data]
        
        tui_to_cui_map, display_name_to_cui_map = get_cui_from_access(tui_values, display_names, connection_main)
        logging.info(f"Completed fetching CUI values for value set {idx}")

        ws = wb.create_sheet(title=str(idx))  # Creating a new sheet with a numbered title
        
        # Adding headers
        headers = ['Description ID', 'DisplayName', 'IncludeChildren', 'Concept ID from Description', 'Concept ID from DisplayName', 'Best Concept ID from Description or DisplayName', 'New Concept ID Exists']
        ws.append(headers)
        
        all_codes_column = set()  # We'll use a set to ensure uniqueness

        # Collect all final_id values in a list
        all_final_ids = []  # Initialize an empty list to collect final_id values

        # Check if any final_id in the current value set has include_children set to "true"
        if not any(entry[2] == "true" for entry in value_set_data):  # Assuming entry[2] is include_children
            logging.info(f"No child codes need to be looked up for value set {idx}.")
        else:
            logging.info(f"Fetching child codes for set {idx} from {transitive_closure_db_path}")
        # Populate all_final_ids list
        for entry in value_set_data:
            value, display_name, include_children, exceptions = entry
            cui_value = tui_to_cui_map.get(value, "Not Found")
            display_name_cui_value = display_name_to_cui_map.get(display_name, "Not Found")
            final_id = cui_value if cui_value != "Not Found" else display_name_cui_value
            all_final_ids.append(final_id)
            checked_cuis.add(final_id)  # Add the final_id to the checked_cuis set

        # Populate new_cui_map using the populated all_final_ids list
        new_cui_map = get_new_cui_from_history(all_final_ids, connection_history)

        # Process each entry in value_set_data and append rows to the worksheet
        for entry in value_set_data:
            value, display_name, include_children, exceptions = entry
            cui_value = tui_to_cui_map.get(value, "Not Found")
            display_name_cui_value = display_name_to_cui_map.get(display_name, "Not Found")
            final_id = cui_value if cui_value != "Not Found" else display_name_cui_value
            new_cui = new_cui_map.get(final_id)
            row = [value, display_name, include_children, cui_value, display_name_cui_value, final_id, new_cui if new_cui else "N/A"]
            ws.append(row)

            # Always add the parent code to all_codes_column, but only if it's not "Not Found"
            if final_id != "Not Found":
                all_codes_column.add(final_id)

            # If IncludeChildren is true and final_id is not "Not Found", traverse the hierarchy and fetch child codes
            if include_children == "true":
                codes_to_check_for_children = [final_id]  # Start with the original code
                if new_cui:  # If there's a new Concept ID, add it to the list of codes to check
                    codes_to_check_for_children.append(new_cui)
                
                for code in codes_to_check_for_children:
                    if code != "Not Found":
                        try:
                            children = get_all_children_from_database(code, connection_tc, exceptions=exceptions)
                            all_codes_column.update(children)
                        except Exception as e:
                            logging.info(f"Did not include child codes for {code}: {e}")
                            continue
        
        # Remove 'Not Found' from the all_codes_column
        all_codes_column = {code for code in all_codes_column if code != "Not Found"}

        # Now write the all_codes_column to Column J
        ws['J1'] = 'All Concepts including Children'  # Setting the header for J
        ws['K1'] = 'Terms for All Concepts'  # Setting the header for K
        
        # Fetch the terms for the codes in all_codes_column
        logging.info(f"Fetching terms for 'All Codes' set {idx}")
       
        # Add all new CUIs to the list of codes
        all_codes_column.update(new_cui_map.values())
        code_to_term_map = {}
        if all_codes_column:
            try:
                codes_placeholder = ",".join(['?'] * len(all_codes_column))
                query = f"SELECT CUI, Term FROM SCT WHERE CUI IN ({codes_placeholder})"
                cursor_main = connection_main.cursor()
                cursor_main.execute(query, list(all_codes_column))
                for row in cursor_main.fetchall():
                    code_to_term_map[row.CUI] = row.Term
            except Exception as e:
                logging.info(f"Exception occurred while fetching terms for 'All Codes' set {idx}: {e}")
                continue
            logging.info(f"Completed fetching terms for 'All Codes' set {idx}\n")

        # Initialize the row index for columns J and K
        jk_row_index = 2

        # Populate columns J and K
        for code, term in code_to_term_map.items():
            if code in all_codes_column:
                ws[f'J{jk_row_index}'] = code
                ws[f'K{jk_row_index}'] = term
                jk_row_index += 1  # Increment the row index

        processed_value_sets = idx  # Update the counter

    if not wb.worksheets:
        logging.info(f"No value sets were processed as they were empty")
        return (0, 0)  # Return processed_value_sets and total_value_sets both as 0

    wb.save(file_path)
    return processed_value_sets, total_value_sets

if __name__ == "__main__":
    
    connection_str_main = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={database_path};'
    )
    connection_main = pyodbc.connect(connection_str_main)

    connection_str_tc = (
    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    f'DBQ={transitive_closure_db_path};'
    )
    connection_tc = pyodbc.connect(connection_str_tc)

    connection_str_history = (
    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    f'DBQ={history_db_path};'
    )
    connection_history = pyodbc.connect(connection_str_history) 

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Extract and process reports from the XML
    for xml_file in os.listdir(xml_directory):
        if xml_file.endswith(".xml"):
            xml_path = os.path.join(xml_directory, xml_file)
            base_name = os.path.basename(xml_path).replace(".xml", "")
            logging.info(f"Starting to process: {xml_file}")
            extract_and_process_reports_from_xml(xml_path, database_path, transitive_closure_db_path, output_dir)
    
    connection_main.close()
    connection_tc.close()
    connection_history.close()
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Script executed in {elapsed_time:.2f} seconds.")
    logging.info(f"Log has been saved to: {log_filename}")