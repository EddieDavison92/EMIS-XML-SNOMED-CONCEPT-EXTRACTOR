import os
import openpyxl
import argparse
import logging
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Set up logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

def format_and_sort_worksheet(worksheet):
    # Sort the data based on columns 'A' and then 'B'
    data = worksheet.values
    headers = next(data, None)
    sorted_data = sorted(data, key=lambda x: (x[0], x[1]))

    worksheet.delete_rows(2, worksheet.max_row - 1)
    for row in sorted_data:
        worksheet.append(row)

    # Apply table style
    tab = Table(displayName="Table1", ref=worksheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    # Adjust column widths
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def consolidate_workbooks(source_dir, output_dir):
    logger.info("Consolidating workbooks from directory: {}".format(source_dir))
    # Ensure the output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Create a new workbook and select the active worksheet
    consolidated_wb = openpyxl.Workbook()
    consolidated_ws = consolidated_wb.active
    consolidated_ws.title = "Consolidated Data"
    consolidated_ws.append(["Workbook", "Sheet", "Code", "Term"])

    # Iterate through each Excel file in the source directory
    for file_name in os.listdir(source_dir):
        if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
            workbook_path = os.path.join(source_dir, file_name)
            wb = openpyxl.load_workbook(workbook_path)
            logger.info(f"Processing workbook: {file_name}")
            # Iterate through each sheet in the workbook
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                # Assuming data starts from row 2, adjust if needed
                for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
                    code, term = row[9], row[10]  # Columns J and K
                    if code is not None and term is not None:
                        consolidated_ws.append([file_name, sheet_name, code, term])
            wb.close()

    # Sort and format the consolidated worksheet
    format_and_sort_worksheet(consolidated_ws)
    
    # Define the filename for the new workbook
    output_file = os.path.join(output_dir, "Consolidated_Workbook.xlsx")

    # Save the new workbook
    consolidated_wb.save(output_file)
    logger.info(f"Consolidated workbook saved to: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Consolidate Workbooks')
    parser.add_argument('--source_dir', type=str, required=True, help='Source directory for workbooks')
    parser.add_argument('--output_dir', type=str, required=True, help='Output directory for the consolidated workbook')

    args = parser.parse_args()
    consolidate_workbooks(args.source_dir, args.output_dir)

if __name__ == '__main__':
    main()

