import xml.etree.ElementTree as ET
import os
import pyodbc
import time
import re
import logging
import argparse
from directory_functions import initialize_directory_structure
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Argument parsing
parser = argparse.ArgumentParser(description="Get paths from inputs")
parser.add_argument("--xml_directory", required=True, help="XML Directory Path")
parser.add_argument("--database_path", required=True, help="Database Path")
parser.add_argument("--transitive_closure_db_path", required=True, help="Transitive Closure DB Path")
parser.add_argument("--history_db_path", required=True, help="History DB Path")
parser.add_argument("--output_dir", required=True, help="Output Directory")

args = parser.parse_args()

# Extract the arguments
xml_directory = args.xml_directory
database_path = args.database_path
transitive_closure_db_path = args.transitive_closure_db_path
history_db_path = args.history_db_path
output_dir = args.output_dir

# Constants
IGNORED_VALUES = ['ACTIVE','REVIEW', 'ENDED', 'N/A', '385432009','C','U','R','RD','999011011000230107','12464001000001103', 'None']
NAMESPACE = {'ns': 'http://www.e-mis.com/emisopen'}

def setup_logger(log_file_path):
    logger = logging.getLogger("main_logger")
    for handler in logger.handlers[:]:
        handler.close()
        logger.removeHandler(handler)

    # Formatter
    formatter = logging.Formatter('%(message)s')
        
    # Stream handler
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    logger.addHandler(stream_handler)
    logger.setLevel(logging.INFO)
    return logger

log_file_path = os.path.join(output_dir, "log.txt")
logger = setup_logger(log_file_path)

start_time = time.time() #start clock

# List XML files
xml_files = [f for f in os.listdir(xml_directory) if f.endswith('.xml')]
logger.info(f"Found {len(xml_files)} XML files in the directory: {xml_directory}")
for file in xml_files:
    logger.info(f"{file}")
logger.info("-" * 40)

def sanitize_filename(filename):
    """Remove characters that are illegal in filenames on Windows."""
    illegal_chars = r'<>:"/\|?*'
    for char in illegal_chars:
        filename = filename.replace(char, '_')
    return filename

def process_single_report(data, report_name, database_path, transitive_closure_db_path, output_dir):
    # Generate the output file path based on the report name
    output_file = os.path.join(output_dir, f"snomed_codes_{report_name}.xlsx")
    
    # Create output folder if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Save the report data to the Excel file
    processed_value_sets, total_value_sets = save_to_xlsx(data, output_file, connection_main, connection_tc, connection_history)
    
    if processed_value_sets > 0:
        logger.info(f"For report '{report_name}', successfully processed {processed_value_sets}/{total_value_sets} value sets.")
        logger.info(f"Excel workbook saved to {output_file}\n")
    else:
        logger.info(f"For report '{report_name}', no value sets were processed as they didn't contain any SNOMED-CT Concepts. No workbook saved.\n")

def extract_and_process_reports_from_xml(xml_path, database_path, transitive_closure_db_path, output_dir):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    reports = root.findall(".//ns:report", NAMESPACE)
    logger.info(f"Found {len(reports)} reports:")

    # Print names of all reports
    report_names = [report.find(".//ns:name", NAMESPACE).text for report in reports]
    for name in report_names:
        logger.info(name)
    logger.info("-" * 40)  # Print separator for clarity

    for report in reports:
        report_name = report.find(".//ns:name", NAMESPACE).text 
        logger.info(f"Processing report: {report_name}")
        name_content = report.find("ns:name", NAMESPACE).text
        match = re.search(r'\[(.*?)\]', name_content)
        report_name_sanitized = sanitize_filename(report_name)  # Sanitize the report name here
        report_name = match.group(1) if match else report_name_sanitized
        extracted_data = extract_values_from_xml_element(report)
        process_single_report(extracted_data, report_name, database_path, transitive_closure_db_path, output_dir)

def extract_values_from_xml_element(element):
    data_sets = []
    seen_data_sets = set()

    for valueSet in element.findall(".//ns:valueSet", NAMESPACE):
        data = []

        for values in valueSet.findall(".//ns:values", NAMESPACE):
            value_elem = values.find('ns:value', NAMESPACE)
            value = value_elem.text if value_elem is not None else "N/A"

            if value in IGNORED_VALUES:
                continue

            displayName_elem = values.find('ns:displayName', NAMESPACE)
            displayName = displayName_elem.text if displayName_elem is not None else "N/A"

            includeChildren_elem = values.find('ns:includeChildren', NAMESPACE)
            includeChildren = includeChildren_elem.text if includeChildren_elem is not None else "false"

            exception_codes = set()
            for exception in valueSet.findall('.//ns:exception//ns:values/ns:value', NAMESPACE):
                exception_codes.add(exception.text) 

            data.append((value, displayName, includeChildren, frozenset(exception_codes)))

        tuple_data = tuple(data)
        if tuple_data not in seen_data_sets and len(data) > 0:
            seen_data_sets.add(tuple_data)
            data_sets.append(data)
    
    logger.info(f"Extracted {len(data_sets)} datasets from the XML.")
    for i, dataset in enumerate(data_sets, 1):
        logger.info(f"Dataset {i}/{len(data_sets)} contains {len(dataset)} value{'s' if len(dataset) != 1 else ''}.")

    return [list(ds) for ds in data_sets]

def get_cui_from_access(tui_list, display_names, connection_main):
    tui_to_cui = {}
    display_name_to_cui = {}
    logger.info(f"Using the established connection to the database")

    # Ensure that the TUI list has distinct values
    distinct_tui_list = list(set(tui_list))

    # Query the database for CUI values based on the provided TUI list
    if distinct_tui_list:
        query_for_tui = "SELECT TUI, CUI FROM SCT WHERE TUI IN ({})".format(','.join(['?'] * len(distinct_tui_list)))
        cursor = connection_main.cursor()
        cursor.execute(query_for_tui, distinct_tui_list)
        tui_to_cui = {row.TUI: row.CUI for row in cursor.fetchall()}
        logger.info(f"Matched {len(tui_to_cui)} Description IDs from XML with SNOMED Description IDs (TUI) from the database and return Concept IDs (CUI).")

    # Query the database for CUI values based on the provided DisplayNames (Terms)
    if display_names: 
        query_for_display_name = "SELECT Term, CUI FROM SCT WHERE Term IN ({})".format(','.join(['?'] * len(display_names)))
        cursor = connection_main.cursor()
        cursor.execute(query_for_display_name, display_names)
        display_name_to_cui = {row.Term: row.CUI for row in cursor.fetchall()}
        logger.info(f"Matched {len(display_name_to_cui)} DisplayNames from XML with SNOMED Description IDs (TUI) from the database and return Concept IDs (CUI).")

    return tui_to_cui, display_name_to_cui

def get_all_children_from_database(code, connection, exceptions=None):
    if code is None:
        return set()
        
    if exceptions is None:
        exceptions = set()
    
    children = {code}
    newly_added = {code}
    
    loop_count = 0

    while newly_added:
        # Increment the loop count
        loop_count += 1

        # Get direct children of the codes in newly_added
        query = "SELECT SubtypeID FROM SCTTC WHERE SupertypeID IN ({})".format(",".join(['?'] * len(newly_added)))
        cursor = connection.cursor()
        cursor.execute(query, list(newly_added))
        new_children = {row.SubtypeID for row in cursor.fetchall()}
        
        # Remove the ones we already know about to avoid infinite loops
        newly_added = new_children - children

        # Exclude exception codes
        excluded_codes = newly_added.intersection(exceptions)
        if excluded_codes:
            logger.info(f"Excluded child codes for {code}: {', '.join(map(str, excluded_codes))}")
        newly_added -= exceptions

        # Add the newly discovered children to main set
        children.update(newly_added)

    if len(children) == 1:  # Only the code itself, no children
        additional_msg = ". No child codes found."
    elif len(children) > 1:  # Code itself plus additional children
        additional_msg = f". Found {len(children) - 1} child codes."
    else: 
        additional_msg = ""

    logger.info(f"Fetching children for code {code} completed in {loop_count} iteration{'s' if loop_count > 1 else ''}{additional_msg}")
    return children

def get_new_cui_from_history(old_cui_list, connection_history):
    if not old_cui_list:  # Check if the list is empty
        return {}
    cursor = connection_history.cursor()
    placeholders = ', '.join('?' for _ in old_cui_list)
    query = f"SELECT OLDCUI, NEWCUI FROM SCTHIST WHERE OLDCUI IN ({placeholders})"
    cursor.execute(query, old_cui_list)
    results = {row.OLDCUI: row.NEWCUI for row in cursor.fetchall()}
    
    # Print details about the history lookups
    for old_cui, new_cui in results.items():
        if new_cui:
            logger.info(f"Old Concept ID: {old_cui} has a new Concept ID: {new_cui}")
        else:
            logger.info(f"No new Concept ID found for: {old_cui}")

    # Ensure that all OLDCUI values are in the results dictionary, with a value of None if no match was found
    for old_cui in old_cui_list:
        results.setdefault(old_cui, None)
    
    return results

def fetch_cui_and_display_maps(value_set_data, connection_main):
    tui_values = [entry[0] for entry in value_set_data]
    display_names = [entry[1] for entry in value_set_data]
    return get_cui_from_access(tui_values, display_names, connection_main)

def process_value_set(value_set_data, ws, connection_main, connection_tc, connection_history, checked_cuis):
    tui_to_cui_map, display_name_to_cui_map = fetch_cui_and_display_maps(value_set_data, connection_main)
    all_codes_column, all_final_ids = set(), []
    
    # Populate all_final_ids here
    for entry in value_set_data:
        value, display_name, include_children, exceptions = entry
        cui_value, display_name_cui_value, final_id = fetch_cui_values(tui_to_cui_map, display_name_to_cui_map, value, display_name)

        if final_id is not None:  # Check to ensure final_id is not None before appending
            all_final_ids.append(final_id)
    
    # Fetch new CUIs based on history after populating all_final_ids
    new_cui_map = get_new_cui_from_history(all_final_ids, connection_history)

    populate_worksheet(ws, value_set_data, tui_to_cui_map, display_name_to_cui_map, new_cui_map, all_codes_column, all_final_ids, connection_tc, checked_cuis)
    write_all_concepts_to_columns(ws, all_codes_column, connection_main)
    return len(all_codes_column)

def populate_worksheet(ws, value_set_data, tui_to_cui_map, display_name_to_cui_map, new_cui_map, all_codes_column, all_final_ids, connection_tc, checked_cuis):
    # Main loop to populate worksheet
    for entry in value_set_data:
        value, display_name, include_children, exceptions = entry
        cui_value, display_name_cui_value, final_id = fetch_cui_values(tui_to_cui_map, display_name_to_cui_map, value, display_name)
        new_cui = new_cui_map.get(final_id)

        if final_id is not None and final_id != "Not Found":  # Check to ensure final_id is not "Not Found" before appending
            all_final_ids.append(final_id)
        
        ws.append([value, display_name, include_children, cui_value, display_name_cui_value, final_id, new_cui if new_cui else "N/A"])
        handle_children_and_update_codes(all_codes_column, final_id, new_cui, include_children, connection_tc, exceptions, checked_cuis)

        if final_id != "Not Found":
            all_final_ids.append(final_id)

def fetch_cui_values(tui_to_cui_map, display_name_to_cui_map, value, display_name):
    cui_value = tui_to_cui_map.get(value, "Not Found")
    display_name_cui_value = display_name_to_cui_map.get(display_name, "Not Found")
    final_id = cui_value if cui_value != "Not Found" else display_name_cui_value
    return cui_value, display_name_cui_value, final_id

def handle_children_and_update_codes(all_codes_column, final_id, new_cui, include_children, connection_tc, exceptions, checked_cuis):
    if final_id != "Not Found":
        all_codes_column.add(final_id)
        checked_cuis.add(final_id)
    
    if include_children == "true":
        # Existing and potential new Concept IDs
        for code in filter(lambda x: x != "Not Found", [final_id, new_cui]):
            try:
                children = get_all_children_from_database(code, connection_tc, exceptions=exceptions)
                all_codes_column.update(children)
            except Exception as e:
                logger.info(f"Did not include child codes for {code}: {e}")

def write_all_concepts_to_columns(ws, all_codes_column, connection_main):
    ws['J1'], ws['K1'] = 'All Concepts including Children', 'Terms for All Concepts'
    code_to_term_map = fetch_all_terms(all_codes_column, connection_main)
    populate_columns_j_and_k(ws, code_to_term_map, all_codes_column)

def fetch_all_terms(all_codes_column, connection_main):
    # Fetch terms for all_codes_column
    code_to_term_map = {}
    try:
        query = f"SELECT CUI, Term FROM SCT WHERE CUI IN ({','.join(['?'] * len(all_codes_column))})"
        cursor_main = connection_main.cursor()
        cursor_main.execute(query, list(all_codes_column))
        code_to_term_map = {row.CUI: row.Term for row in cursor_main.fetchall()}
    except Exception as e:
        logger.info(f"Exception while fetching terms: {e}")
    return code_to_term_map

def populate_columns_j_and_k(ws, code_to_term_map, all_codes_column):
    for idx, (code, term) in enumerate(code_to_term_map.items(), start=2):
        if code in all_codes_column:
            ws[f'J{idx}'] = code
            ws[f'K{idx}'] = term

def save_to_xlsx(data, file_path, connection_main, connection_tc, connection_history):
    total_value_sets = len(data)
    wb = Workbook()
    wb.remove(wb.active) 
    
    checked_cuis = set()
    for idx, value_set_data in enumerate(data, 1):
        ws = wb.create_sheet(title=str(idx))
        ws.append(['Description ID', 'DisplayName', 'IncludeChildren', 'Concept ID from Description', 'Concept ID from DisplayName', 'Best Concept ID from Description or DisplayName', 'New Concept ID Exists'])
        
        process_value_set(value_set_data, ws, connection_main, connection_tc, connection_history, checked_cuis)
    
    if not wb.worksheets:
        return 0, 0
    wb.save(file_path)
    return len(wb.worksheets), total_value_sets

if __name__ == "__main__":
    
    connection_str_main = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={database_path};'
    )
    connection_main = pyodbc.connect(connection_str_main)

    connection_str_tc = (
    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    f'DBQ={transitive_closure_db_path};'
    )
    connection_tc = pyodbc.connect(connection_str_tc)

    connection_str_history = (
    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    f'DBQ={history_db_path};'
    )
    connection_history = pyodbc.connect(connection_str_history) 

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Extract and process reports from the XML
    for xml_file in os.listdir(xml_directory):
        if xml_file.endswith(".xml"):
            xml_path = os.path.join(xml_directory, xml_file)
            base_name = os.path.basename(xml_path).replace(".xml", "")
            logger.info(f"Starting to process: {xml_file}")
            extract_and_process_reports_from_xml(xml_path, database_path, transitive_closure_db_path, output_dir)
    
    connection_main.close()
    connection_tc.close()
    connection_history.close()
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"Script executed in {elapsed_time:.2f} seconds.")
    logger.info(f"Log has been saved to: {output_dir}")